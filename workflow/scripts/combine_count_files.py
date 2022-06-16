from xlsxwriter.workbook import Workbook
import os
import openpyxl
import argparse

# Initialize parser
parser = argparse.ArgumentParser()
# Adding optional argument
parser.add_argument("-o", "--Output", help="Show Output")
parser.add_argument("-i1", "--Input1", help="Show Input species_count")
parser.add_argument("-i2", "--Input2", help="Show Input amr_count")
parser.add_argument("-i3", "--Input3", help="Show Input count_match")
# Read arguments from command line
args = parser.parse_args()
i1 = args.Input1
i2 = args.Input2
i3 = args.Input3
target = args.Output

if i1 and i2 and i3 and target:
    # path = r'database/test/count_files/'
    # target = r'database/test/combine_count_files_confer.xlsx'
    file_list = [i1, i2, i3]

    # Create an XlsxWriter workbook object and add a worksheet.
    workbook = Workbook(target)
    worksheet = workbook.add_worksheet()

    # Add a bold format to use to highlight cells.
    bold = workbook.add_format({'bold': True})

    worksheet.write('A1', 'Name', bold)
    worksheet.write('B1', 'Number of hits', bold)
    worksheet.write('D1', 'Name', bold)
    worksheet.write('E1', 'Number of hits', bold)
    worksheet.write('G1', 'Name', bold)
    worksheet.write('H1', 'Number of hits', bold)

    # Start from the first cell. Rows and columns are zero indexed.
    row = 1
    col = 0
    # row and collumn file
    file_row = 1
    file_col = 1

    for file in file_list:
        if file.endswith('species_count.xlsx'):
            # Define variable to load the wookbook
            wookbook = openpyxl.load_workbook(file)
            # Define variable to read the active sheet:
            sheet = wookbook.active
            col = 0
            species_last = sheet.max_row
            for row in range(1, sheet.max_row+1): #elke rij van een count file wordt afgegaan en de geschreven in de gesamenvatten file
                name = sheet.cell(row=file_row, column=file_col).value #eerste kolom van een count file is de naam van gen/species
                hit = sheet.cell(row=file_row, column=file_col + 1).value #tweede kolom van een count file is het aantal hits
                worksheet.write(row, col, name) #schrijven van naam gen/species
                worksheet.write(row, col + 1, hit) #schrijven van hit
                row += 1 #naar de volgende rij om te schrijven
                file_row += 1 #naar de volgende rij om te lezen
            row = 1
            file_row = 1
        if file.endswith('amr_count.xlsx'):
            # Define variable to load the wookbook
            wookbook = openpyxl.load_workbook(file)
            # Define variable to read the active sheet:
            sheet = wookbook.active
            col = 3
            amr_last = sheet.max_row
            for row in range(1, sheet.max_row+1):  # elke rij van een count file wordt afgegaan en de geschreven in de gesamenvatten file
                name = sheet.cell(row=file_row, column=file_col).value  # eerste kolom van een count file is de naam van gen/species
                hit = sheet.cell(row=file_row, column=file_col + 1).value  # tweede kolom van een count file is het aantal hits
                worksheet.write(row, col, name)  # schrijven van naam gen/species
                worksheet.write(row, col + 1, hit)  # schrijven van hit
                row += 1  # naar de volgende rij om te schrijven
                file_row += 1  # naar de volgende rij om te lezen
            row = 1
            file_row = 1
        if file.endswith('count_match.xlsx'):
            # Define variable to load the wookbook
            wookbook = openpyxl.load_workbook(file)
            # Define variable to read the active sheet:
            sheet = wookbook.active
            col = 6
            match_last = sheet.max_row
            for row in range(1, sheet.max_row+1):  # elke rij van een count file wordt afgegaan en de geschreven in de gesamenvatten file
                name = sheet.cell(row=file_row, column=file_col).value  # eerste kolom van een count file is de naam van gen/species
                hit = sheet.cell(row=file_row, column=file_col + 1).value  # tweede kolom van een count file is het aantal hits
                worksheet.write(row, col, name)  # schrijven van naam gen/species
                worksheet.write(row, col + 1, hit)  # schrijven van hit
                row += 1  # naar de volgende rij om te schrijven
                file_row += 1  # naar de volgende rij om te lezen
            row = 1
            file_row = 1

    # Create a new chart object.
    chart1 = workbook.add_chart({'type': 'pie'})

    # Configure the series. Note the use of the list syntax to define ranges:
    #     [sheetname, first_row, first_col, last_row, last_col]
    chart1.add_series({
        'name':       'species data chart',
        'categories': ['Sheet1', 1, 0, species_last, 0],
        'values':     ['Sheet1', 1, 1, species_last, 1],
        'data_labels': {'value': True, 'category': True, 'percentage': True, 'legend_key': False},
    })

    # Add a title.
    chart1.set_title({'name': 'Species'})
    # Set an Excel chart style. Colors with white outline and shadow.
    chart1.set_style(10)
    # Remove legend
    chart1.set_legend({'none': True})
    # Set chart size
    chart1.set_size({'x_scale': 2, 'y_scale': 2})
    # Insert the chart into the worksheet.
    worksheet.insert_chart('J1', chart1)

    ##############################################################################

    # Create a new chart object.
    chart2 = workbook.add_chart({'type': 'pie'})

    # Configure the series. Note the use of the list syntax to define ranges:
    #     [sheetname, first_row, first_col, last_row, last_col]
    chart2.add_series({
        'name':       'amr data chart',
        'categories': ['Sheet1', 1, 3, amr_last, 3],
        'values':     ['Sheet1', 1, 4, amr_last, 4],
        'data_labels': {'value': True, 'category': True, 'percentage': True, 'legend_key': False},
    })

    # Add a title.
    chart2.set_title({'name': 'Antibiotic Resistance Genes'})
    # Set an Excel chart style. Colors with white outline and shadow.
    chart2.set_style(10)
    # Remove legend
    chart2.set_legend({'none': True})
    # Set chart size
    chart2.set_size({'x_scale': 2, 'y_scale': 2})
    # Insert the chart into the worksheet.
    worksheet.insert_chart('J31', chart2)

    ##############################################################################

    # Create a new chart object.
    chart3 = workbook.add_chart({'type': 'pie'})

    # Configure the series. Note the use of the list syntax to define ranges:
    #     [sheetname, first_row, first_col, last_row, last_col]
    chart3.add_series({
        'name':       'amr_species data chart',
        'categories': ['Sheet1', 1, 6, match_last, 6],
        'values':     ['Sheet1', 1, 7, match_last, 7],
        'data_labels': {'value': True, 'category': True, 'percentage': True, 'legend_key': False},
    })

    # Add a title.
    chart3.set_title({'name': 'AMR + Species Matches'})
    # Set an Excel chart style. Colors with white outline and shadow.
    chart3.set_style(10)
    # Remove legend
    chart3.set_legend({'none': True})
    # Set chart size
    chart3.set_size({'x_scale': 2, 'y_scale': 2})
    # Insert the chart into the worksheet.
    worksheet.insert_chart('J61', chart3)

    workbook.close()