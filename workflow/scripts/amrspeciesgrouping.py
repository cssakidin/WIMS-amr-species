import openpyxl
import argparse
from xlsxwriter.workbook import Workbook

# Initialize parser
parser = argparse.ArgumentParser()
# Adding optional argument
parser.add_argument("-o", "--Output", help="Show Output")
parser.add_argument("-i", "--Input", help="Show Input species_count")
# Read arguments from command line
args = parser.parse_args()
target = args.Output

if args.Input and target:
    # Create an XlsxWriter workbook object and add a worksheet.
    workbook = Workbook(target)
    worksheet = workbook.add_worksheet()

    with open(args.Input) as cmb_f:
        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook(file)
        # Define variable to read the active sheet:
        sheet = wookbook.active
        species_last = sheet.max_row
        for row in range(1, sheet.max_row+1): #elke rij van een count file wordt afgegaan
            name = sheet.cell(row=file_row, column=file_col).value  # eerste kolom van een count file is de naam van gen/species
            hit = sheet.cell(row=file_row, column=file_col + 1).value  # tweede kolom van een count file is het aantal hits
            if name not in dic:
